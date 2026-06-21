"""
Plot Results + Export to Excel
Generates comparison figures and Excel with:
  - Column 1: Time (s)
  - Column 2: p99 Latency (s)
  - Column 3: Number of CPU Cores (replicas)
  Plus summary comparing all 3 experiments
"""

import os
import glob
import pandas as pd
import matplotlib.pyplot as plt

RESULTS_DIR = "results"
OUTPUT_DIR  = "results/figures"

LABEL_MAP = {
    "custom_autoscaler": "Custom Autoscaler",
    "hpa_70":            "HPA 70% CPU",
    "hpa_90":            "HPA 90% CPU",
}
COLORS = {
    "custom_autoscaler": "#1f77b4",
    "hpa_70":            "#ff7f0e",
    "hpa_90":            "#2ca02c",
}


def load_csvs(results_dir: str) -> dict:
    """Load all experiment CSV files and normalize their time axis."""
    dfs = {}
    for path in sorted(glob.glob(os.path.join(results_dir, "*.csv"))):
        df = pd.read_csv(path)
        if df.empty or "experiment" not in df.columns:
            continue
        label = df["experiment"].iloc[0]
        if "timestamp" in df.columns:
            df["time_s"] = df["timestamp"] - df["timestamp"].iloc[0]
        dfs[label] = df
    return dfs


def plot_combined(dfs: dict):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 9), sharex=True)

    for label, df in dfs.items():
        color = COLORS.get(label)
        name  = LABEL_MAP.get(label, label)

        if "p99" in df.columns:
            s = df["p99"].rolling(3, min_periods=1).mean()
            ax1.plot(df["time_s"], s, label=name, color=color, linewidth=2)

        if "ready_replicas" in df.columns:
            ax2.plot(df["time_s"], df["ready_replicas"],
                     label=name, color=color, linewidth=2)

    ax1.axhline(0.5, color="red", linestyle="--", linewidth=1.5, label="SLO (0.5s)")
    ax1.set_ylabel("p99 Latency (s)", fontsize=12)
    ax1.legend(fontsize=10)
    ax1.grid(True, alpha=0.3)
    ax1.set_title("Autoscaler Comparison: p99 Latency & CPU Cores", fontsize=14, fontweight="bold")

    ax2.set_ylabel("Number of CPU Cores", fontsize=12)
    ax2.set_xlabel("Time (seconds)", fontsize=12)
    ax2.legend(fontsize=10)
    ax2.grid(True, alpha=0.3)

    plt.tight_layout()
    out = os.path.join(OUTPUT_DIR, "comparison_combined.png")
    plt.savefig(out, dpi=150)
    print(f"Saved plot: {out}")
    plt.close()


def export_excel(dfs: dict):
    """Export each experiment and a summary sheet into a styled Excel workbook."""
    try:
        import openpyxl
    except ImportError:
        os.system("pip install openpyxl -q")

    os.makedirs(RESULTS_DIR, exist_ok=True)
    outfile = os.path.join(RESULTS_DIR, "results_summary.xlsx")

    with pd.ExcelWriter(outfile, engine="openpyxl") as writer:

        # ── One sheet per experiment ──────────────────────────────────────────
        for label, df in dfs.items():
            sheet = LABEL_MAP.get(label, label)[:31]

            cols = {
                "time_s":          "Time (s)",
                "p99":             "p99 Latency (s)",
                "ready_replicas":  "Number of CPU Cores",
                "p50":             "p50 Latency (s)",
                "mean":            "Mean Latency (s)",
                "count":           "Request Count",
                "slo_violations":  "SLO Violations",
                "slo_rate":        "SLO Violation Rate",
            }
            available = {k: v for k, v in cols.items() if k in df.columns}
            df_out = df[list(available.keys())].copy()
            df_out.columns = list(available.values())
            df_out["Time (s)"] = df_out["Time (s)"].round(1)
            df_out.to_excel(writer, sheet_name=sheet, index=False)

            # Style header
            from openpyxl.styles import Font, PatternFill, Alignment
            ws = writer.sheets[sheet]
            for cell in ws[1]:
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = PatternFill("solid", fgColor="1F4E79")
                cell.alignment = Alignment(horizontal="center")

            # Highlight SLO violations
            red_fill = PatternFill("solid", fgColor="FFCCCC")
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                val = row[1].value  # p99 column
                if val and float(val) > 0.5:
                    for cell in row:
                        cell.fill = red_fill

            # Auto-size columns
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col) + 3
                ws.column_dimensions[col[0].column_letter].width = min(max_len, 25)

        # ── Summary sheet ─────────────────────────────────────────────────────
        summary = []
        for label, df in dfs.items():
            p99  = df["p99"].dropna()
            reps = df["ready_replicas"].dropna()
            summary.append({
                "Experiment":                  LABEL_MAP.get(label, label),
                "Avg p99 Latency (s)":         round(float(p99.mean()), 3),
                "Max p99 Latency (s)":         round(float(p99.max()), 3),
                "Min p99 Latency (s)":         round(float(p99.min()), 3),
                "Total SLO Violations":        int(df["slo_violations"].sum()) if "slo_violations" in df else "-",
                "CPU Cores at Start":          int(reps.iloc[0]),
                "CPU Cores at End":            int(reps.iloc[-1]),
                "Max CPU Cores":               int(reps.max()),
                "Duration (s)":                round(float(df["time_s"].max()), 0),
            })

        df_summary = pd.DataFrame(summary)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)

        from openpyxl.styles import Font, PatternFill, Alignment
        ws = writer.sheets["Summary"]
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col) + 3
            ws.column_dimensions[col[0].column_letter].width = min(max_len, 30)

        # ── Charts sheet ──────────────────────────────────────────────────────
        chart_path = os.path.join(OUTPUT_DIR, "comparison_combined.png")
        if os.path.exists(chart_path):
            from openpyxl.drawing.image import Image as XLImage
            wb  = writer.book
            ws3 = wb.create_sheet("Charts")
            ws3["A1"] = "Autoscaler Comparison: p99 Latency & CPU Cores"
            ws3["A1"].font = Font(bold=True, size=12)
            img = XLImage(chart_path)
            img.width  = 800
            img.height = 550
            ws3.add_image(img, "A2")

    print(f"Saved Excel: {outfile}")
    print(f"  Sheets: {[LABEL_MAP.get(l,l) for l in dfs]} + Summary + Charts")


def main():
    dfs = load_csvs(RESULTS_DIR)
    if not dfs:
        print(f"No experiment CSVs found in {RESULTS_DIR}")
        return

    print(f"Loaded: {list(dfs.keys())}")
    plot_combined(dfs)
    export_excel(dfs)

    print("\nDone!")
    print(f"  Plot:  {OUTPUT_DIR}/comparison_combined.png")
    print(f"  Excel: {RESULTS_DIR}/results_summary.xlsx")


if __name__ == "__main__":
    main()
